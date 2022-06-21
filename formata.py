from numpy import var
import openpyxl
import att_xlsx as a

bookFor = openpyxl.Workbook()


bookFor.create_sheet('Janeiro')
bookFor.create_sheet('Fevereiro')
bookFor.create_sheet('Março')
bookFor.create_sheet('Abril')
bookFor.create_sheet('Maio')
bookFor.create_sheet('Junho')
bookFor.create_sheet('Julho')
bookFor.create_sheet('Agosto')
bookFor.create_sheet('Setembro')
bookFor.create_sheet('Outubro')
bookFor.create_sheet('Novembro')
bookFor.create_sheet('Dezembro')

#print(bookFor.sheetnames)

bookFor.remove(bookFor['Sheet'])

book2 = openpyxl.load_workbook('Planilha de Acs-Certs.xlsx')

# arrayGeral = []
# arrayNome = []
# arrayDeleta = []

# for sheets in book2:
#     print(sheets)
#     for rows in sheets:
#         arrayGeral.append(rows[0].value)
#         for elements in arrayGeral:
#             if elements not in arrayNome:
#                 arrayNome.append(rows[0].value)
        
# print(arrayNome)

dict = {}
lista = []
for sheets in book2:
    for rows in sheets:
        lista += [rows[0].value]
        

print(set(lista))


bookFor.save('Total_AC_Mes.xlsx')        