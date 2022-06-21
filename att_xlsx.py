#Comando no cmd - baixar biblioteca:
# pip install openpyxl OU pip3 install openpyxl

import openpyxl
from os import path, walk, listdir
from ntpath import join
import zipfile
from datetime import datetime

book = openpyxl.Workbook()


book.create_sheet('Janeiro')
book.create_sheet('Fevereiro')
book.create_sheet('Março')
book.create_sheet('Abril')
book.create_sheet('Maio')
book.create_sheet('Junho')
book.create_sheet('Julho')
book.create_sheet('Agosto')
book.create_sheet('Setembro')
book.create_sheet('Outubro')
book.create_sheet('Novembro')
book.create_sheet('Dezembro')

#print(book.sheetnames)

book.remove(book['Sheet'])

#print(book.sheetnames)

def achar_mes(ano, mes, contador):
    if zips.startswith(ano):
        certs = book[mes]
        novo_zip = zips.replace(' ', '_').replace('-','_')[9:-4].upper()
        certs.append([novo_zip, contador])
    


path_zips = r'M:/teste_zips_certs'

for zips in listdir(path_zips):
    caminho = path.join(path_zips, zips)
    contador_certificados = 0
    # print(zips)
    try:
        with zipfile.ZipFile(caminho, 'r') as myzip:
            for zip in myzip.namelist():
                if zip.endswith('.cer'):
                    contador_certificados += 1
        
            achar_mes('202101', 'Janeiro', contador_certificados)
            achar_mes('202102', 'Fevereiro', contador_certificados)
            achar_mes('202103', 'Março', contador_certificados)
            achar_mes('202104', 'Abril', contador_certificados)
            achar_mes('202105', 'Maio', contador_certificados)
            achar_mes('202106', 'Junho', contador_certificados)
            achar_mes('202107', 'Julho', contador_certificados)
            achar_mes('202108', 'Agosto', contador_certificados)
            achar_mes('202109', 'Setembro', contador_certificados)
            achar_mes('202110', 'Outubro', contador_certificados)
            achar_mes('202111', 'Novembro', contador_certificados)
            achar_mes('202112', 'Dezembro', contador_certificados)


            book.save('Planilha de Acs-Certs.xlsx')

    except Exception as e:
        print(e)
        try:
            date = datetime.now().strftime("%Y%m%d")
            logname = (date+"_"+"Log_Errado"+".txt")
            mensagem = (zips + ", está corrompido.")
        # Gerar a mensagem
            with open(logname, 'a', encoding='utf-8') as log:
                log.write(f'{mensagem}\n')
        except Exception as e:
            print(e, 'Erro ao criar o Log_Errado')
            pass

    # tup = (zips, contador_certificados)
    # print(tup)
